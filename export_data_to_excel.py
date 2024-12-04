
import os
import json
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def main():
    file_name = 'my_new_book.xlsx'
    abs_path = os.path.abspath(file_name)
    if os.path.isfile(abs_path):
        print('YES')

        with open('result_zhana_urpaq.json') as file:
            data = json.load(file)

        book = load_workbook(r'my_new_book.xlsx')
        sheet = book.active

        # Установить основные заголовки
        sheet['A3'] = 'APT_NUMBER'
        sheet['B3'] = 'ROOMS_Q'
        sheet['C3'] = 'APT_S'
        sheet['D3'] = 'FLOOR'
        sheet['E3'] = 'ENTRANCE'
        sheet['G3'] = 'DATE'
        sheet['H3'] = 'PRICE_SQ_M'
        sheet['I3'] = 'PRICE_TOTAL'

        row = 4
        for item in data:
            apt_number = item['Номер квартиры:']
            price_sq_m_new = int(item['Цена за квм:'].replace(' ', ''))
            price_total_new = price_sq_m_new * float(item['Площадь:'])

            # Поиск следующей свободной группы колонок справа
            col_offset = 0
            while sheet.cell(row=row, column=7 + col_offset).value is not None:
                col_offset += 3  # Сдвиг на 3 колонки

            # Добавление заголовков для новых колонок
            if sheet.cell(row=3, column=7 + col_offset).value is None:
                # Записываем новые данные в следующую свободную группу колонок
                sheet.cell(row=row, column=7 + col_offset).value = datetime.date.today()  # Дата
                sheet.cell(row=row, column=8 + col_offset).value = price_sq_m_new  # PRICE_SQ_M
                sheet.cell(row=row, column=9 + col_offset).value = price_total_new  # PRICE_TOTAL

                # Проверка изменений и закраска только для PRICE_SQ_M
                if col_offset > 0:  # Если это не первая группа данных
                    prev_price_sq_m = sheet.cell(row=row, column=8 + col_offset - 3).value

                    if prev_price_sq_m is not None and price_sq_m_new is not None:
                        # Приведение к float для корректного сравнения
                        if float(price_sq_m_new) < float(prev_price_sq_m):
                            fill_color = PatternFill(start_color="00FF00", end_color="00FF00",
                                                     fill_type="solid")  # Зеленый
                        elif float(price_sq_m_new) > float(prev_price_sq_m):
                            fill_color = PatternFill(start_color="FF0000", end_color="FF0000",
                                                     fill_type="solid")  # Красный
                        else:
                            fill_color = None

                        # Применяем цвет, если он задан
                        if fill_color:
                            sheet.cell(row=row, column=8 + col_offset).fill = fill_color

            # Записываем новые данные в следующую свободную группу колонок
            sheet.cell(row=row, column=7 + col_offset).value = datetime.date.today()  # Дата
            sheet.cell(row=row, column=8 + col_offset).value = price_sq_m_new  # PRICE_SQ_M
            sheet.cell(row=row, column=9 + col_offset).value = price_total_new  # PRICE_TOTAL

            # Заполняем остальные данные, если не заполнены
            sheet.cell(row=row, column=1).value = apt_number
            sheet.cell(row=row, column=2).value = item['Количество комнат:']
            sheet.cell(row=row, column=3).value = item['Площадь:']
            sheet.cell(row=row, column=4).value = item['Этаж:']
            sheet.cell(row=row, column=5).value = item['Подъезд:']

            row += 1

        book.save(r'my_new_book.xlsx')
        book.close()

    else:
        print('FILE NOT FOUND')
        print('Creating file ...')

        with open('result_zhana_urpaq.json') as file:
            data = json.load(file)

        book = Workbook()
        sheet = book.active

        # Устанавливаем основные заголовки
        sheet['A3'] = 'APT_NUMBER'
        sheet['B3'] = 'ROOMS_Q'
        sheet['C3'] = 'APT_S'
        sheet['D3'] = 'FLOOR'
        sheet['E3'] = 'ENTRANCE'
        sheet['G3'] = 'DATE'
        sheet['H3'] = 'PRICE_SQ_M'
        sheet['I3'] = 'PRICE_TOTAL'

        row = 4
        for item in data:
            apt_number = item['Номер квартиры:']
            price_sq_m_new = int(item['Цена за квм:'].replace(' ', ''))
            price_total_new = price_sq_m_new * float(item['Площадь:'])

            # Добавление заголовков для первой группы данных
            if sheet.cell(row=3, column=7).value is None:
                sheet.cell(row=3, column=7).value = 'DATE'
                sheet.cell(row=3, column=8).value = 'PRICE_SQ_M'
                sheet.cell(row=3, column=9).value = 'PRICE_TOTAL'

            # Заполняем данные
            sheet.cell(row=row, column=1).value = apt_number
            sheet.cell(row=row, column=2).value = item['Количество комнат:']
            sheet.cell(row=row, column=3).value = item['Площадь:']
            sheet.cell(row=row, column=4).value = item['Этаж:']
            sheet.cell(row=row, column=5).value = item['Подъезд:']
            sheet.cell(row=row, column=7).value = datetime.date.today()
            sheet.cell(row=row, column=8).value = price_sq_m_new
            sheet.cell(row=row, column=9).value = price_total_new

            row += 1

        book.save('my_new_book.xlsx')
        book.close()

    print('done')

if __name__ == "__main__":
    main()
