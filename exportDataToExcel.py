import os
import json
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def make_excel(project_name):
    print(f'project_name: {project_name}'
          f'\n\nPAGE: DOS_exportDataToExcel.py - make_excel')

    file_name = f'{project_name}.xlsx'
    abs_path = os.path.abspath(file_name)
    json_name = f'{project_name}.json'

    if os.path.isfile(abs_path):
        print('YES')

        # Загружаем данные из JSON
        with open(f'{json_name}') as file:
            data = json.load(file)

        # Загружаем Excel-файл
        book = load_workbook(fr'{file_name}')
        sheet = book.active

        # Сохраняем обработанные номера квартир
        processed_apartments = set()

        for item in data:
            apt_number = item['Номер квартиры:']
            price_sq_m_new = int(item['Цена за квм:'].replace(' ', ''))
            price_total_new = price_sq_m_new * float(item['Площадь:'])

            # Поиск строки квартиры
            row = 4
            while sheet.cell(row=row, column=1).value is not None:
                if sheet.cell(row=row, column=1).value == apt_number:
                    break
                row += 1
            else:
                # Если квартира отсутствует, добавляем новую строку
                sheet.cell(row=row, column=1).value = apt_number
                sheet.cell(row=row, column=2).value = item['Количество комнат:']
                sheet.cell(row=row, column=3).value = item['Площадь:']
                sheet.cell(row=row, column=4).value = item['Этаж:']
                sheet.cell(row=row, column=5).value = item['Подъезд:']

            processed_apartments.add(apt_number)

            # Поиск следующей свободной группы колонок справа
            col_offset = 0
            while sheet.cell(row=row, column=7 + col_offset).value is not None:
                col_offset += 3

            # Добавление новых данных
            sheet.cell(row=row, column=7 + col_offset).value = datetime.date.today()  # Дата
            sheet.cell(row=row, column=8 + col_offset).value = price_sq_m_new  # PRICE_SQ_M
            sheet.cell(row=row, column=9 + col_offset).value = price_total_new  # PRICE_TOTAL

            # Проверка изменений цены и закраска ячейки PRICE_SQ_M
            if col_offset > 0:  # Если есть предыдущие данные
                prev_price_sq_m = sheet.cell(row=row, column=8 + col_offset - 3).value

                if prev_price_sq_m is not None and price_sq_m_new != float(prev_price_sq_m):
                    if price_sq_m_new < float(prev_price_sq_m):
                        fill_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Зеленый
                    elif price_sq_m_new > float(prev_price_sq_m):
                        fill_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Красный
                    else:
                        fill_color = None

                    if fill_color:
                        sheet.cell(row=row, column=8 + col_offset).fill = fill_color

        # Квартиры, которые не обработаны (удаленные из JSON), остаются без изменений

        book.save(fr'{file_name}')
        book.close()

    else:
        print('FILE NOT FOUND')
        print('Creating file ...')

        with open(f'{json_name}') as file:
            data = json.load(file)

        book = Workbook()
        sheet = book.active

        # Устанавливаем заголовки
        sheet['A3'] = 'APT_NUMBER'
        sheet['B3'] = 'ROOMS_Q'
        sheet['C3'] = 'APT_S'
        sheet['D3'] = 'FLOOR'
        sheet['E3'] = 'ENTRANCE'
        sheet['G3'] = 'DATE'
        sheet['H3'] = 'PRICE_SQ_M'
        sheet['I3'] = 'PRICE_TOTAL'

        row = 4
        for item in data:
            apt_number = item['Номер квартиры:']
            price_sq_m_new = int(item['Цена за квм:'].replace(' ', ''))
            price_total_new = price_sq_m_new * float(item['Площадь:'])

            sheet.cell(row=row, column=1).value = apt_number
            sheet.cell(row=row, column=2).value = item['Количество комнат:']
            sheet.cell(row=row, column=3).value = item['Площадь:']
            sheet.cell(row=row, column=4).value = item['Этаж:']
            sheet.cell(row=row, column=5).value = item['Подъезд:']
            sheet.cell(row=row, column=7).value = datetime.date.today()
            sheet.cell(row=row, column=8).value = price_sq_m_new
            sheet.cell(row=row, column=9).value = price_total_new

            row += 1

        book.save(f'{file_name}')
        book.close()

    print('done')

